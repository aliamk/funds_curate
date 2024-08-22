import streamlit as st
import pandas as pd
import tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Utility Functions
def copy_columns(source_df, mapping, additional_values=None):
    dest_df = pd.DataFrame()
    for source_col, dest_col in mapping.items():
        if source_col in source_df.columns:
            dest_df[dest_col] = source_df[source_col]
        else:
            dest_df[dest_col] = additional_values.get(dest_col, "") if additional_values else ""
    return dest_df

def append_close_data(writer, source_df, status, event_type, title_suffix, startrow, report):
    close_filter = source_df["STATUS"] == status
    close_funds = source_df[close_filter]
    if close_funds.empty:
        report.append(f"{status} data => data was not found in the source file\n")
        return startrow

    close_df = pd.DataFrame({
        "Fund": close_funds["NAME"],
        "Event Date": close_funds["LATEST INTERIM CLOSE DATE"],
        "Event Type": event_type,
        "Title": close_funds["NAME"] + title_suffix,
        "Close Size": close_funds["LATEST INTERIM CLOSE SIZE (CURR. MN)"]
    })

    if not close_df.empty:
        close_df.to_excel(writer, sheet_name='Events', index=False, header=False, startrow=startrow)
        report.append(f"{status} data => from row {startrow + 2}\n")
        startrow += len(close_df)

    return startrow

def append_performance_data(writer, source_df, startrow, report):
    performance_mapping = {
        "NAME": "Fund",
        "": "Performance Date",
        "TARGET IRR - GROSS MIN": "Performance Value (Min)",
        "TARGET IRR - GROSS MAX": "Performance Value (Max)",
        "": "Fund Performance Measurement Type",
        "": "Fund Performance Measurement Unit",
        "": "Performance Source",
        "": "Confidential"
    }
    
    additional_values = {
        "Fund Performance Measurement Unit": "Percentage",
        "Fund Performance Measurement Type": "Target IRR (Gross) (%)"
    }
    
    performance_df = copy_columns(source_df, performance_mapping, additional_values)
    
    performance_df['Performance Source'] = ""
    performance_df['Confidential'] = ""

    # Ensure the correct order of columns and insertion if necessary
    if 'Fund Performance Measurement Type' not in performance_df.columns:
        performance_df.insert(2, 'Fund Performance Measurement Type', "Target IRR (Gross) (%)")
    else:
        performance_df['Fund Performance Measurement Type'] = "Target IRR (Gross) (%)"
    
    if 'Fund Performance Measurement Unit' not in performance_df.columns:
        performance_df.insert(3, 'Fund Performance Measurement Unit', "Percentage")
    else:
        performance_df['Fund Performance Measurement Unit'] = "Percentage"

    if not performance_df.empty:
        performance_df.to_excel(writer, sheet_name='Performances', index=False, header=False, startrow=startrow)
        report.append(f"Target IRR (Gross) (%) data => from row {startrow + 1}\n")
        startrow += len(performance_df)
    
    return startrow

def autofit_columns(writer_path):
    workbook = load_workbook(writer_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.row == 1:  # Check if the cell is in the header row
                    cell.alignment = Alignment(horizontal='left')  # Align header left
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
    workbook.save(writer_path)

def append_roles(writer, source_df, role_column, role_name, startrow, report):
    roles_data = {
        "Fund": source_df["NAME"],
        "Company": source_df.get(role_column, ""),
        "Role": role_name
    }
    roles_df_add = pd.DataFrame(roles_data)
    roles_df_add.to_excel(writer, sheet_name='Roles', index=False, header=False, startrow=startrow)
    report.append(f"{role_name} data => from row {startrow + 1}\n")
    return startrow + len(roles_df_add)

def record_replacement(report, original, replacement, count, rows):
    rows_str = ', '.join(map(str, rows[:5])) + (' ...' if len(rows) > 5 else '')
    report.append(f"'{original}': '{replacement}' => {count} replacements (rows {rows_str})")

# Tab Creation Functions
def create_funds_tab(writer, source_df, report):
    funds_mapping = {
        "NAME": "Fund",
        "FUND CURRENCY": "Fund Currency",
        "VINTAGE / INCEPTION YEAR": "Vintage Year",
        "STATUS": "Fund Status",
        "STRATEGY": "Fund Style",
        "ASSET CLASS": "Asset Class",
        "FUND STRUCTURE": "Separate Account",
        "FUND NUMBER (OVERALL)": "Fund Sequence (Total)",
        "FUND NUMBER (SERIES)": "Fund Series",
        "LIFESPAN (YEARS)": "Fund Life",
        "LIFESPAN EXTENSION": "Fund Life Extension",
        "TARGET SIZE (CURR. MN)": "Target Size (Local Currency m)",
        "INITIAL TARGET (CURR. MN)": "Initial Target Size (Local Currency m)",
        "HARD CAP (CURR. MN)": "Hard Cap (Local Currency m)",
        "OFFER CO-INVESTMENT OPPORTUNITIES TO LPS?": "Fund coinvesting Lps",
        "FUND LEGAL STRUCTURE": "Fund Legal Structure",
        "TIMES TO FIRST CLOSE": "Times to First Close",
        "TOTAL MONTHS IN MARKET": "Total Months in Market",
        "OVERIDE FUND STATUS": "Overide Fund Status"
    }
    
    funds_df = copy_columns(source_df, funds_mapping)
    
    # Apply Fund Status and Fund Style rules
    fund_status_replacements = {
        'Open Ended': 'Open ended',
        'Open ended (Liquidated)': 'Open ended',
        'Semi-Open Ended': 'Quasi-open ended',
        'Evergreen': 'Evergreen',
        'Closed': 'Final Close',
        'Open-Ended (Liquidated)': 'Liquidated',
        'Semi-Open Ended': 'Quasi-open ended',
        'Raising': 'Launched',
        'Estimated': 'Speculative',
        'Listed': '',
    }
    funds_df['Open/Closed'] = funds_df.apply(lambda row: (
        'Open ended' if row['Fund Status'] in ['Open Ended', 'Open ended (Liquidated)'] else
        'Quasi-open ended' if row['Fund Status'] == 'Semi-Open Ended' else
        'Evergreen' if row['Fund Status'] == 'Evergreen' else
        'Closed ended'
    ), axis=1)

    replacements_notes = {}
    for original, replacement in fund_status_replacements.items():
        count = funds_df['Fund Status'].value_counts().get(original, 0)
        if count > 0:
            rows = funds_df.index[funds_df['Fund Status'] == original].tolist()
            replacements_notes[original] = (replacement, count, rows)
        funds_df['Fund Status'] = funds_df['Fund Status'].replace(original, replacement)

    # Fund Style replacements
    fund_style_replacements = {
        'Value Added': 'Value Add',
        'Debt': 'Debt',
        'Opportunistic': 'Opportunistic',
        'Core-Plus': 'Core Plus',
        'Real Asset': '',
        'Core': 'Core',
        'Distressed': 'Distressed',
        'Fund of Funds': 'Fund of Funds',
        'Co-Investment': 'PE Co-Investment',
        'Real Asset Fund of Funds': 'Fund of Funds',
        'Secondaries': 'Secondaries',
        'Infrastructure Core': 'Core',
        'Mezzanine': 'Mezzanine',
        'Hybrid': 'Hybrid',
        'Venture (General)': 'VC',
        'CMBS': '',
        'Hybrid Fund of Funds': 'Fund of Funds',
        'Direct Lending': 'Debt',
        'Credit/Securities': '',
        'Real Estate CMBS': '',
        'Real Estate Core': 'Core',
        'Real Estate Core-Plus': 'Core Plus',
        'Real Estate Debt': 'Debt',
        'Real Estate Distressed': 'Distressed',
        'Real Estate Fund of Funds': 'Fund of Funds',
        'Real Estate Opportunistic': 'Opportunistic',
        'Real Estate Value Added': 'Value Add',
        'Infrastructure Core Plus': 'Core Plus',
        'Infrastructure Opportunistic': 'Opportunistic',
        'Infrastructure Value Added': 'Value Add',
        'Infrastructure Fund of Funds': 'Fund of Funds',
        'Infrastructure Debt': 'Debt',
        'Infrastructure Secondaries': 'Secondaries',
        'Core Plus': 'Core Plus',
        'Real Estate Secondaries': 'Secondaries',
        'Real Estate Co-Investment': '',
        'Infrastructure': '',
        'Buyout': 'PE Buyout',
        'Co-Investment Multi-Manager': 'PE Co-Investment',
        'Direct Lending - Blended / Opportunistic Debt': 'Direct Lending',
        'Direct Lending - Senior Debt': 'Direct Lending',
        'Distressed Debt': 'Distressed',
        'Early Stage: Start-up': 'VC Early Stage',
        'Expansion / Late Stage': 'VC Late Stage',
        'Growth': 'PE Growth',
    }
    for original, replacement in fund_style_replacements.items():
        count = funds_df['Fund Style'].value_counts().get(original, 0)
        if count > 0:
            rows = funds_df.index[funds_df['Fund Style'] == original].tolist()
            replacements_notes[original] = (replacement, count, rows)
        funds_df['Fund Style'] = funds_df['Fund Style'].replace(original, replacement)

    # Asset Class replacements
    asset_class_replacements = {
        'Multi': 'Diversified',
    }
    for original, replacement in asset_class_replacements.items():
        count = funds_df['Asset Class'].value_counts().get(original, 0)
        if count > 0:
            rows = funds_df.index[funds_df['Asset Class'] == original].tolist()
            replacements_notes[original] = (replacement, count, rows)
        funds_df['Asset Class'] = funds_df['Asset Class'].replace(original, replacement)

    # Separate Account replacements
    separate_account_replacements = {
        'Commingled': 'No',
        'Separately Managed Account': 'Yes',
    }
    for original, replacement in separate_account_replacements.items():
        count = funds_df['Separate Account'].value_counts().get(original, 0)
        if count > 0:
            rows = funds_df.index[funds_df['Separate Account'] == original].tolist()
            replacements_notes[original] = (replacement, count, rows)
        funds_df['Separate Account'] = funds_df['Separate Account'].replace(original, replacement)

    # Fund Life Extension replacements
    funds_df['Fund Life Extension'] = funds_df['Fund Life Extension'].astype(str).str.replace(r'\+', ';', regex=True)
    funds_df['Fund Life Extension'] = funds_df['Fund Life Extension'].str.replace('nan', '', regex=False)

    # Override Fund Status
    override_replacements = funds_df['Fund Status'] == 'Liquidated'
    override_count = override_replacements.sum()
    override_rows = funds_df.index[override_replacements].tolist()
    funds_df.loc[override_replacements, 'Overide Fund Status'] = "True"

    # Reorder columns to match the desired order
    column_order = [
        "Fund", "Fund Currency", "Vintage Year", "Fund Status", "Open/Closed", "Fund Style", "Asset Class",
        "Separate Account", "Fund Legal Structure", "Fund Sequence (Total)", "Fund Series", "Fund Life", 
        "Fund Life Extension", "Target Size (Local Currency m)", "Initial Target Size (Local Currency m)",
        "Hard Cap (Local Currency m)", "Fund coinvesting Lps", "Times to First Close", "Total Months in Market",
        "Overide Fund Status"
    ]
    funds_df = funds_df[column_order]

    funds_df.to_excel(writer, sheet_name='Funds', index=False)
    report.append("Funds tab created")
    report.append(f"{len(funds_df.columns)} Columns\n")
    report.extend(funds_df.columns)
    report.append("\n///////////////////////////////////////////////////////////////////////////\n")

    report.append("'Funds' tab adjustments\n")
    for original, (replacement, count, rows) in replacements_notes.items():
        record_replacement(report, original, replacement, count, rows)
    report.append(f"'Liquidated': 'True' => {override_count} replacements (rows {', '.join(map(str, override_rows[:5])) + (' ...' if len(override_rows) > 5 else '')})")
    report.append("\n///////////////////////////////////////////////////////////////////////////\n")

def create_events_tab(writer, source_df, report):
    # Check if 'FINAL CLOSE DATE' and other necessary columns exist in the DataFrame
    if 'FINAL CLOSE DATE' in source_df.columns and 'FINAL CLOSE SIZE (CURR. MN)' in source_df.columns:
        final_close_date = source_df["FINAL CLOSE DATE"]

        events_mapping = {
            "NAME": "Fund",
            "FUND RAISING LAUNCH DATE": "Event Date",
            "": "Event Type",
            "": "Title",
            "FINAL CLOSE SIZE (CURR. MN)": "Close Size"
        }
        
        events_df = copy_columns(source_df, events_mapping)
        events_df['Event Type'] = "Launch"
        events_df['Title'] = events_df['Fund'] + " launches"

        # Remove rows with blank Event Date
        events_df = events_df[events_df['Event Date'].notna() & (events_df['Event Date'] != '')]

        events_df.to_excel(writer, sheet_name='Events', index=False)
        report.append("Events tab created")
        report.append(f"{len(events_df.columns)} Columns\n")
        report.extend(events_df.columns)
        report.append("\nLaunch data entered from row 2\n")

        # Append additional data to Events tab
        additional_data = {
            "Fund": source_df["NAME"],
            "Event Date": final_close_date,
            "Event Type": "Final Close",
            "Title": source_df["NAME"] + " reaches final close",
            "Close Size": source_df["FINAL CLOSE SIZE (CURR. MN)"]
        }
        additional_df = pd.DataFrame(additional_data)
        additional_df = additional_df[additional_df['Event Date'].notna()]
        additional_df.to_excel(writer, sheet_name='Events', index=False, header=False, startrow=len(events_df)+1)
        report.append(f"Final Close data => from row {len(events_df) + 2}\n")
        startrow = len(events_df) + len(additional_df) + 1

        # Append data for various closes
        startrow = append_close_data(writer, source_df, "First Close", "First Close", " reaches first close", startrow, report)
        startrow = append_close_data(writer, source_df, "Second Close", "Second Close", " reaches second close", startrow, report)
        startrow = append_close_data(writer, source_df, "Third Close", "Third Close", " reaches third close", startrow, report)
        startrow = append_close_data(writer, source_df, "Fourth Close", "Fourth Close", " reaches fourth close", startrow, report)
        startrow = append_close_data(writer, source_df, "Fifth Close", "Fifth Close", " reaches fifth close", startrow, report)
        startrow = append_close_data(writer, source_df, "Sixth Close", "Sixth Close", " reaches sixth close", startrow, report)
        startrow = append_close_data(writer, source_df, "Seventh Close", "Seventh Close", " reaches seventh close", startrow, report)
        report.append("///////////////////////////////////////////////////////////////////////////\n")
    else:
        report.append("Required columns for 'FINAL CLOSE DATE' or 'FINAL CLOSE SIZE (CURR. MN)' not found. Skipping final close data.\n")

def create_performances_tab(writer, source_df, report):
    performances_mapping = {
        "NAME": "Fund",
        "": "Performance Date",
        "": "Fund Performance Measurement Type",
        "": "Fund Performance Measurement Unit",
        "TARGET IRR - NET MIN": "Performance Value (Min)",
        "TARGET IRR - NET MAX": "Performance Value (Max)",
        "": "Performance Source",
        "": "Confidential"
    }
    additional_values_performances = {
        "Fund Performance Measurement Unit": "Percentage",
        "Fund Performance Measurement Type": "Target IRR Net"
    }
    performances_df = copy_columns(source_df, performances_mapping, additional_values_performances)
    for col in ['Fund', 'Performance Date', 'Fund Performance Measurement Type', 'Fund Performance Measurement Unit', 'Performance Value (Min)', 'Performance Value (Max)', 'Performance Source', 'Confidential']:
        if col not in performances_df.columns:
            performances_df[col] = ""
    performances_df['Performance Date'] = ""
    performances_df['Fund Performance Measurement Type'] = "Target IRR Net"
    performances_df['Fund Performance Measurement Unit'] = "Percentage"
    performances_df['Performance Source'] = ""
    performances_df['Confidential'] = ""
    performances_df = performances_df[['Fund', 'Performance Date', 'Fund Performance Measurement Type', 'Fund Performance Measurement Unit', 'Performance Value (Min)', 'Performance Value (Max)', 'Performance Source', 'Confidential']]
    
    # Remove rows where both Performance Value (Min) and Performance Value (Max) are blank    
    performances_df = performances_df[~((performances_df['Performance Value (Min)'].isna()) & (performances_df['Performance Value (Max)'].isna()) | 
                                  ((performances_df['Performance Value (Min)'] == '') & (performances_df['Performance Value (Max)'] == '')))]

    performances_df.to_excel(writer, sheet_name='Performances', index=False)
    report.append("Performances tab created")
    report.append(f"{len(performances_df.columns)} Columns\n")
    report.extend(performances_df.columns)

    startrow_performance = len(performances_df) + 1
    startrow_performance = append_performance_data(writer, source_df, startrow_performance, report)
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_actual_performance_tab(writer, report):
    actual_performance_df = pd.DataFrame(columns=["Fund", "Performance Date", "Called (%)"])
    actual_performance_df.to_excel(writer, sheet_name='Actual Performance', index=False)
    report.append("Actual Performance tab created")
    report.append("3 Columns\n")
    report.extend(["Fund", "Performance Date", "Called (%)"])
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_domicile_tab(writer, source_df, report):
    domicile_mapping = {
        "NAME": "Fund",
        "DOMICILE": "Domicile"
    }
    domicile_df = copy_columns(source_df, domicile_mapping)
    
    # Domicile replacements
    domicile_replacements = {
        'Alberta': 'United States',
        'Arizona': 'United States',
        'Australia': 'Australia',
        'Bahrain': 'Bahrain',
        'Belgium': 'Belgium',
        'Bermuda': 'Bermuda',
        'Brazil': 'Brazil',
        'British Virgin Islands': 'British Virgin Islands',
        'California': 'United States',
        'Canada': 'Canada',
        'Cayman Islands': 'Cayman Islands',
        'Chile': 'Chile',
        'China': 'China',
        'Colombia': 'Colombia',
        'Colorado': 'Colorado',
        'Cyprus': 'Cyprus',
        'Czech Republic': 'Czech Republic',
        'Delaware': 'United States',
        'Denmark': 'Denmark',
        'England': 'United Kingdom',
        'Estonia': 'Estonia',
        'Finland': 'Finland',
        'Florida': 'Florida',
        'France': 'France',
        'Georgia': 'United States',
        'Germany': 'Germany',
        'Guernsey': 'Guernsey',
        'Hungary': 'Hungary',
        'Illinois': 'United States',
        'India': 'India',
        'Ireland': 'Ireland',
        'Italy': 'Italy',
        'Japan': 'Japan',
        'Jersey': 'Jersey',
        'Kansas': 'United States',
        'Kuwait': 'Kuwait',
        'Lithuania': 'Lithuania',
        'Louisiana': 'United States',
        'Luxembourg': 'Luxembourg',
        'Malaysia': 'Malaysia',
        'Maryland': 'United States',
        'Massachusetts': 'United States',
        'Mauritius': 'Mauritius',
        'Mexico': 'Mexico',
        'Michigan': 'United States',
        'Minnesota': 'United States',
        'Missouri': 'United States',
        'Morocco': 'Morocco',
        'Nebraska': 'United States',
        'Netherlands': 'Netherlands',
        'Nevada': 'United States',
        'New Jersey': 'United States',
        'New York': 'United States',
        'New Zealand': 'New Zealand',
        'North Carolina': 'United States',
        'North Dakota': 'United States',
        'Norway': 'Norway',
        'Ohio': 'United States',
        'Oklahoma': 'United States',
        'Ontario': 'Ontario',
        'Oregon': 'United States',
        'Pennsylvania': 'United States',
        'Peru': 'Peru',
        'Poland': 'Poland',
        'Portugal': 'Portugal',
        'Romania': 'Romania',
        'Russia': 'Russia',
        'Saudi Arabia': 'Saudi Arabia',
        'Singapore': 'Singapore',
        'Slovenia': 'Slovenia',
        'South Africa': 'South Africa',
        'South Carolina': 'South Carolina',
        'South Dakota': 'United States',
        'South Korea': 'South Korea',
        'Spain': 'Spain',
        'St. Lucia': 'St. Lucia',
        'Sweden': 'Sweden',
        'Switzerland': 'Switzerland',
        'Tennessee': 'United States',
        'Texas': 'United States',
        'UK': 'United Kingdom',
        'Ukraine': 'Ukraine',
        'United Kingdom': 'United Kingdom',
        'US': 'United States',
        'Utah': 'United States',
        'Virginia': 'United States',
        'Washington': 'United States',
        'Wisconsin': 'United States',
        'Wyoming': 'United States',
        'Latvia': 'Latvia',
        'Kenya': 'Kenya',
        'Taiwan': 'Taiwan',
        'Malta': 'Malta',
        'Panama': 'Panama',
        'EU': '',
        'Hong Kong': 'Hong Kong',
        'Isle of Man': '',
        'United Arab Emirates': 'United Arab Emirates',
        'Liechtenstein': 'Liechtenstein',
        'Maine': '',
        'Greece': 'Greece',
        'Israel': 'Israel',
        'Indonesia': 'Indonesia',
        'Nigeria': 'Nigeria',
        'Marshall Islands': '',
        'Scotland': 'United Kingdom',
    }
    
    domicile_notes = {}
    for original, replacement in domicile_replacements.items():
        count = domicile_df['Domicile'].value_counts().get(original, 0)
        if count > 0:
            rows = domicile_df.index[domicile_df['Domicile'] == original].tolist()
            domicile_notes[original] = (replacement, count, rows)
        domicile_df['Domicile'] = domicile_df['Domicile'].replace(original, replacement)
    
    domicile_df.to_excel(writer, sheet_name='Domicile', index=False)
    report.append("Domicile tab created")
    report.append(f"{len(domicile_df.columns)} Columns\n")
    report.extend(domicile_df.columns)
    
    report.append("'Domicile' tab adjustments\n")
    for original, (replacement, count, rows) in domicile_notes.items():
        record_replacement(report, original, replacement, count, rows)
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_target_geographies_primary_region_tab(writer, source_df, report):
    target_geographies_primary_region_mapping = {
        "NAME": "Fund",
        "PRIMARY REGION FOCUS": "Target Geographies Primary Region"
    }
    target_geographies_primary_region_df = copy_columns(source_df, target_geographies_primary_region_mapping)
    
    # Target Geographies Primary Region replacements
    primary_region_replacements = {
        'Diversified Multi-Regional': 'Multi-Region',
        'Americas': 'North America, Latin America & Caribbean',
        'Middle East & Israel': 'Middle East & North Africa',
        'Africa': 'Sub-Saharan Africa',
    }
    
    primary_region_notes = {}
    for original, replacement in primary_region_replacements.items():
        count = target_geographies_primary_region_df['Target Geographies Primary Region'].value_counts().get(original, 0)
        if count > 0:
            rows = target_geographies_primary_region_df.index[target_geographies_primary_region_df['Target Geographies Primary Region'] == original].tolist()
            primary_region_notes[original] = (replacement, count, rows)
        target_geographies_primary_region_df['Target Geographies Primary Region'] = target_geographies_primary_region_df['Target Geographies Primary Region'].replace(original, replacement)
    
    target_geographies_primary_region_df.to_excel(writer, sheet_name='Target_Geographies_Primary_Regi', index=False)
    report.append("Target_Geographies_Primary_Regi tab created")
    report.append(f"{len(target_geographies_primary_region_df.columns)} Columns\n")
    report.extend(target_geographies_primary_region_df.columns)
    
    report.append("'Target_Geographies_Primary_Regi' tab adjustments\n")
    for original, (replacement, count, rows) in primary_region_notes.items():
        record_replacement(report, original, replacement, count, rows)
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_target_geographies_tab(writer, source_df, report):
    target_geographies_mapping = {
        "NAME": "Fund",
        "GEOGRAPHIC EXPOSURE": "Fund Target Geography"
    }
    target_geographies_df = copy_columns(source_df, target_geographies_mapping)

    
    def replace_geographies(geographies, replacements):
        if pd.isna(geographies):
            return geographies
        geographies_list = [geo.strip() for geo in geographies.split(',')]
        updated_list = [replacements.get(geo, geo) for geo in geographies_list]
        return ', '.join(updated_list)

    # Target Geographies replacements
    geographies_replacements = {
        'Africa': 'Sub-Saharan Africa',
        'Americas': 'North America, Latin America & Caribbean',
        'ASEAN': 'Asia',
        'Asia and Rest of World': 'Multi-Region',
        'Central and East Europe': 'Central & Eastern Europe',
        'East and Southeast Asia': 'East & Southeast Asia',
        'Emerging Markets': 'Multi-Region',
        'EU': 'Europe',
        'Greater China': 'China',
        'Hong Kong SAR - China': 'Hong Kong',
        'Macao SAR - China': 'Macao',
        'Middle East': 'Middle East & North Africa',
        'Nordic': 'Nordics',
        'OECD': 'Multi-Region',
        'South America': 'Latin America & Caribbean',
        'UK': 'United Kingdom',
        'US': 'United States',
        'West Europe': 'Western Europe',
        'MENA': 'Middle East & North Africa',
        'GCC': 'Bahrain, Kuwait, Oman, Qatar, Saudi Arabia, United Arab Emirates',
        'Frontier Markets': 'Multi-Region',
    }

    target_geographies_df['Fund Target Geography'] = target_geographies_df['Fund Target Geography'].apply(replace_geographies, replacements=geographies_replacements)

    
    geographies_notes = {}
    for original, replacement in geographies_replacements.items():
        count = target_geographies_df['Fund Target Geography'].value_counts().get(original, 0)
        if count > 0:
            rows = target_geographies_df.index[target_geographies_df['Fund Target Geography'] == original].tolist()
            geographies_notes[original] = (replacement, count, rows)
        target_geographies_df['Fund Target Geography'] = target_geographies_df['Fund Target Geography'].replace(original, replacement)
    
    target_geographies_df.to_excel(writer, sheet_name='Target_Geographies', index=False)
    report.append("Target_Geographies tab created")
    report.append(f"{len(target_geographies_df.columns)} Columns\n")
    report.extend(target_geographies_df.columns)
    
    report.append("'Target_Geographies' tab adjustments\n")
    for original, (replacement, count, rows) in geographies_notes.items():
        record_replacement(report, original, replacement, count, rows)
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_target_sectors_primary_tab(writer, source_df, report):
    target_sectors_primary_mapping = {
        "NAME": "Fund",
        "INF: PRIMARY SECTOR": "Sector - Primary"
    }
    target_sectors_primary_df = copy_columns(source_df, target_sectors_primary_mapping)
    
    # Target Sectors Primary replacements
    sectors_primary_replacements = {
        'Niche': '',
        'Hotels': 'Hospitality',
        'Operating Companies': '',
        'Hotel': 'Hospitality',
        'Social': 'Social Infrastructure',
        'Energy': 'Oil & Gas',
        'Telecommunications': 'Digital Infrastructure',
        'Waste Management': 'Waste',
        'Utilities': 'Conventional Energy',
        'Hotel': 'Hospitality',
    }
    
    sectors_primary_notes = {}
    for original, replacement in sectors_primary_replacements.items():
        count = target_sectors_primary_df['Sector - Primary'].value_counts().get(original, 0)
        if count > 0:
            rows = target_sectors_primary_df.index[target_sectors_primary_df['Sector - Primary'] == original].tolist()
            sectors_primary_notes[original] = (replacement, count, rows)
        target_sectors_primary_df['Sector - Primary'] = target_sectors_primary_df['Sector - Primary'].replace(original, replacement)
    
    target_sectors_primary_df.to_excel(writer, sheet_name='Target_Sectors_Primary', index=False)
    report.append("Target_Sectors_Primary tab created")
    report.append(f"{len(target_sectors_primary_df.columns)} Columns\n")
    report.extend(target_sectors_primary_df.columns)
    
    report.append("'Target_Sectors_Primary' tab adjustments\n")
    for original, (replacement, count, rows) in sectors_primary_notes.items():
        record_replacement(report, original, replacement, count, rows)
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_target_sectors_secondary_tab(writer, report):
    target_sectors_secondary_df = pd.DataFrame(columns=["Fund", "Fund Subsectors"])
    target_sectors_secondary_df.to_excel(writer, sheet_name='Target_Sectors_Secondary', index=False)
    report.append("Target_Sectors_Secondary tab created")
    report.append("2 Columns\n")
    report.extend(["Fund", "Fund Subsectors"])
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_roles_tab(writer, source_df, report):
    roles_data = []
    
    roles = [
        ("FUND MANAGER", "General Partner"),
        ("PLACEMENT AGENTS", "Placement Agent"),
        ("LAW FIRMS", "Legal Advisor"),
        ("AUDITORS", "Auditor"),
        ("ADMINISTRATORS", "Administrator")
    ]
    
    for role_column, role_name in roles:
        role_df = pd.DataFrame({
            "Fund": source_df["NAME"],
            "Company": source_df.get(role_column, ""),
            "Role": role_name
        })
        roles_data.append(role_df)
    
    roles_df = pd.concat(roles_data, ignore_index=True)
    
    # Normalize the 'Company' column to handle case insensitivity only for specific words
    def normalize_company(value):
        words = str(value).split()
        if words and (words[0].lower() in ['used', 'not']):
            return value.lower()
        return value
    
    roles_df['Company'] = roles_df['Company'].apply(normalize_company)

    # Handle Company replacements and additional columns
    roles_df['Confidential'] = roles_df['Company'].apply(lambda x: "TRUE" if x == "used but not specified" else "")
    used_not_specified_count = roles_df['Company'].value_counts().get('used but not specified', 0)
    roles_df['Company'] = roles_df['Company'].replace("used but not specified", "")
    
    # Add 'Not Used' column and set values
    roles_df['Not Used'] = roles_df['Company'].apply(lambda x: "TRUE" if x == "not used" else "")
    
    # Reorder columns to match the specified order
    roles_df = roles_df[['Fund', 'Company', 'Role', 'Not Used', 'Confidential']]
    
    # Delete rows where 'Company' is blank or contains a comma
    initial_count = len(roles_df)
    roles_df = roles_df[roles_df['Company'].notna()]
    blank_or_comma_deleted = initial_count - len(roles_df)
    initial_count = len(roles_df)
    # Ensure 'Company' column is completely filled and converted to string type
    roles_df['Company'] = roles_df['Company'].fillna('').astype(str)

    # Now apply the filter to remove rows where 'Company' contains a comma
    contains_comma = roles_df['Company'].str.contains(',')

    # Use the ~ operator on the boolean Series to filter out rows with commas
    roles_df = roles_df[~contains_comma]
    comma_deleted = initial_count - len(roles_df)
    
    # Tally for 'Not Used' replacements
    not_used_count = roles_df['Company'].value_counts().get('not used', 0)
    
    # Delete the content for specific values
    roles_df['Company'] = roles_df['Company'].replace("not used", "")
    
    roles_df.to_excel(writer, sheet_name='Roles', index=False)
    report.append("Roles tab created")
    report.append(f"{len(roles_df.columns)} Columns\n")
    report.extend(['Fund', 'Company', 'Role', 'Not Used', 'Confidential'])
    report.append("///////////////////////////////////////////////////////////////////////////\n")
    
    # Update the report with the tallies
    report.append(f"'Roles' tab, column 'Company' rows deleted containing blank/contains commas: {blank_or_comma_deleted}")
    report.append(f"'Roles' tab, column 'Company' rows deleted containing 'Not Used': {not_used_count}")
    report.append(f"'Roles' tab, column 'Company' rows deleted containing 'Used but Not Specified': {used_not_specified_count}")
    report.append("///////////////////////////////////////////////////////////////////////////\n")

def create_fees_tab(writer):
    fees_df = pd.DataFrame(columns=["Fund", "Attribute", "Value"])
    fees_df.to_excel(writer, sheet_name='Fees', index=False)
    
# Main Processing Function
# Utility Function to clean specific values
def clean_specific_values(df, tab_name, report):
    for col in df.columns:
        for idx, value in df[col].items():
            if value in ["0", "nan", "n/a"]:
                report.append(f'"{value}" deleted, "{tab_name}" tab, column "{col}", row {idx + 2}')
                df.at[idx, col] = ""
    return df

# In your process function, apply the clean_specific_values function to each dataframe
def process_file(uploaded_file):
    xls = pd.ExcelFile(uploaded_file)
    sheets = xls.sheet_names
    report = []

    default_sheets = ["Sheet1", "Sheet2", "Sheet3"]  # Replace with your actual sheet names

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            writer.book.create_sheet('Dummy')
            has_data = False
            for sheet in default_sheets:
                if sheet in sheets:
                    source_df = pd.read_excel(xls, sheet_name=sheet)
                    
                    # Clean specific values in the dataframe
                    source_df = clean_specific_values(source_df, sheet, report)
                    
                    create_funds_tab(writer, source_df, report)
                    create_events_tab(writer, source_df, report)
                    create_performances_tab(writer, source_df, report)
                    create_actual_performance_tab(writer, report)
                    create_domicile_tab(writer, source_df, report)
                    create_target_geographies_primary_region_tab(writer, source_df, report)
                    create_target_geographies_tab(writer, source_df, report)
                    create_target_sectors_primary_tab(writer, source_df, report)
                    create_target_sectors_secondary_tab(writer, report)
                    create_roles_tab(writer, source_df, report)
                    create_fees_tab(writer)
                    has_data = True

            if has_data:
                del writer.book['Dummy']

        # Auto-fit column widths
        autofit_columns(tmp.name)

        # Generate file name with current date and time
        now = datetime.now().strftime("%Y%m%d_%H%M")
        dest_file_name = f"curated_finfra1_{now}.xlsx"
        dest_file_path = tmp.name

        # Create report file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.txt') as report_file:
            report_file_path = report_file.name
            report_file.write("\n".join(report).encode('utf-8'))

    return dest_file_path, dest_file_name, report_file_path


# Streamlit UI
st.title("Curating FINFRA 1 data files")
st.write("Upload your source Excel file to create a destination file based on predefined instructions.")

uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

if uploaded_file:
    if 'processed' not in st.session_state:
        with st.spinner('Processing file...'):
            dest_file_path, dest_file_name, report_file_path = process_file(uploaded_file)
            st.session_state.processed = True
            st.session_state.dest_file_path = dest_file_path
            st.session_state.dest_file_name = dest_file_name
            st.session_state.report_file_path = report_file_path

    st.success(f"Destination file '{st.session_state.dest_file_name}' created successfully!")
    
    st.download_button(
        label="Download Destination File",
        data=open(st.session_state.dest_file_path, "rb"),
        file_name=st.session_state.dest_file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.download_button(
        label="Download Report File",
        data=open(st.session_state.report_file_path, "rb"),
        file_name="curated_finfra1_report.txt",
        mime="text/plain"
    )
